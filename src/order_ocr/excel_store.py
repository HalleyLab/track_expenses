from __future__ import annotations

import shutil
from copy import copy
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

from .config import DEFAULT_SHEET, DEFAULT_WORKBOOK
from .parser import OrderCandidate, ReferenceData, normalize_key


HEADER_ALIASES = {
    "item": {"item", "items", "product", "name", "\u7269\u54c1", "\u5546\u54c1"},
    "category": {"category", "type", "\u7c7b\u522b", "\u5206\u7c7b"},
    "purchase_date": {"purchase date", "date", "order date", "\u8d2d\u4e70\u65e5\u671f", "\u4e0b\u5355\u65e5\u671f"},
    "ended": {"ended", "is ended", "\u662f\u5426\u7ed3\u675f"},
    "end_date": {"end date", "\u7ed3\u675f\u65e5\u671f"},
    "duration": {"duration", "days", "\u6301\u7eed\u5929\u6570"},
    "price": {"price", "amount", "\u4ef7\u683c", "\u91d1\u989d"},
    "daily_price": {"daily price", "price per day", "\u65e5\u4ef7"},
}


@dataclass
class AppendResult:
    row: int
    workbook_path: Path
    backup_path: Path | None = None


class BalanceWorkbook:
    def __init__(self, path: Path | str = DEFAULT_WORKBOOK, sheet_name: str = DEFAULT_SHEET):
        self.path = Path(path)
        self.sheet_name = sheet_name

    def load_reference(self) -> ReferenceData:
        wb = load_workbook(self.path, data_only=True, read_only=True)
        try:
            ws = wb[self.sheet_name]
            headers = self._headers_from_sheet(ws)
            item_col = headers.get("item", 1)
            category_col = headers.get("category", 2)

            categories: list[str] = []
            category_seen: set[str] = set()
            item_category: dict[str, str] = {}

            for row in ws.iter_rows(min_row=2, values_only=True):
                item = row[item_col - 1] if len(row) >= item_col else None
                category = row[category_col - 1] if len(row) >= category_col else None
                if not category:
                    continue
                category_text = str(category).strip()
                key = normalize_key(category_text)
                if key and key not in category_seen:
                    categories.append(category_text)
                    category_seen.add(key)
                if item:
                    item_category[str(item).strip()] = category_text
            return ReferenceData(categories=categories, item_category=item_category)
        finally:
            wb.close()

    def append_order(self, order: OrderCandidate, backup: bool = True) -> AppendResult:
        self._validate_order(order)
        backup_path = self._backup() if backup else None

        wb = load_workbook(self.path)
        try:
            ws = wb[self.sheet_name]
            headers = self._headers_from_sheet(ws)
            required = ["item", "category", "purchase_date", "end_date", "duration", "price", "daily_price"]
            missing = [name for name in required if name not in headers]
            if missing:
                raise ValueError(f"Missing workbook headers: {', '.join(missing)}")

            row = ws.max_row + 1
            previous_row = max(row - 1, 2)
            for col in range(1, 9):
                source = ws.cell(previous_row, col)
                target = ws.cell(row, col)
                if source.has_style:
                    target._style = copy(source._style)
                if source.number_format:
                    target.number_format = source.number_format
                if source.alignment:
                    target.alignment = copy(source.alignment)

            item_col = headers["item"]
            category_col = headers["category"]
            purchase_col = headers["purchase_date"]
            ended_col = headers.get("ended")
            end_col = headers["end_date"]
            duration_col = headers["duration"]
            price_col = headers["price"]
            daily_col = headers["daily_price"]

            ws.cell(row, item_col).value = order.item
            ws.cell(row, category_col).value = order.category
            ws.cell(row, purchase_col).value = _as_datetime(order.purchase_date)
            ws.cell(row, purchase_col).number_format = "yyyy-mm-dd"
            if ended_col:
                ws.cell(row, ended_col).value = None
            ws.cell(row, end_col).value = "=TODAY()"
            ws.cell(row, end_col).number_format = "yyyy-mm-dd"

            purchase_ref = f"{get_column_letter(purchase_col)}{row}"
            end_ref = f"{get_column_letter(end_col)}{row}"
            duration_ref = f"{get_column_letter(duration_col)}{row}"
            price_ref = f"{get_column_letter(price_col)}{row}"

            ws.cell(row, duration_col).value = f'=DATEDIF({purchase_ref},{end_ref},"D")+1'
            ws.cell(row, price_col).value = float(order.price)
            ws.cell(row, price_col).number_format = "0.00"
            ws.cell(row, daily_col).value = f"={price_ref}/{duration_ref}"
            ws.cell(row, daily_col).number_format = "0.00"

            wb.save(self.path)
            return AppendResult(row=row, workbook_path=self.path, backup_path=backup_path)
        finally:
            wb.close()

    def _backup(self) -> Path:
        timestamp = datetime.now().strftime("%Y%m%d-%H%M%S")
        backup_path = self.path.with_name(f"{self.path.stem}.backup-{timestamp}{self.path.suffix}")
        shutil.copy2(self.path, backup_path)
        return backup_path

    @staticmethod
    def _headers_from_sheet(ws) -> dict[str, int]:
        raw_headers = [cell.value for cell in ws[1]]
        mapping: dict[str, int] = {}
        for index, value in enumerate(raw_headers, start=1):
            if value is None:
                continue
            key = normalize_key(str(value))
            for canonical, aliases in HEADER_ALIASES.items():
                if key in {normalize_key(alias) for alias in aliases} and canonical not in mapping:
                    mapping[canonical] = index
        return mapping

    @staticmethod
    def _validate_order(order: OrderCandidate) -> None:
        if not order.item:
            raise ValueError("Item is required.")
        if not order.category:
            raise ValueError("Category is required.")
        if not order.purchase_date:
            raise ValueError("Purchase date is required.")
        if order.price is None:
            raise ValueError("Price is required.")


def _as_datetime(value: date | datetime | None) -> datetime | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value
    return datetime(value.year, value.month, value.day)


def create_sample_workbook(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = DEFAULT_SHEET
    ws.append(
        [
            "\u7269\u54c1",
            "\u7c7b\u522b",
            "\u8d2d\u4e70\u65e5\u671f",
            "\u662f\u5426\u7ed3\u675f",
            "\u7ed3\u675f\u65e5\u671f",
            "\u6301\u7eed\u5929\u6570",
            "\u4ef7\u683c",
            "\u65e5\u4ef7",
        ]
    )
    wb.save(path)
