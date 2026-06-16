from datetime import date

from openpyxl import load_workbook

from order_ocr.excel_store import BalanceWorkbook, create_sample_workbook
from order_ocr.parser import OrderCandidate


def test_append_order_writes_expected_formulas(tmp_path):
    workbook_path = tmp_path / "Balance.xlsx"
    create_sample_workbook(workbook_path)

    store = BalanceWorkbook(workbook_path)
    result = store.append_order(
        OrderCandidate(
            item="Dove Body Wash",
            category="Daily Necessities",
            purchase_date=date(2026, 6, 15),
            price=5.99,
        ),
        backup=False,
    )

    wb = load_workbook(workbook_path, data_only=False)
    ws = wb["Sheet1"]

    assert result.row == 2
    assert ws["A2"].value == "Dove Body Wash"
    assert ws["B2"].value == "Daily Necessities"
    assert ws["E2"].value == "=TODAY()"
    assert ws["F2"].value == '=DATEDIF(C2,E2,"D")+1'
    assert ws["G2"].value == 5.99
    assert ws["H2"].value == "=G2/F2"
